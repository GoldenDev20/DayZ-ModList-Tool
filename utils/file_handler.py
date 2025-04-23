import os
import json
import openpyxl
from openpyxl.styles import PatternFill
from jinja2 import Template

def load_config(config_file):
    """Load configuration from a JSON file."""
    with open(config_file, "r") as f:
        config = json.load(f)
    return config

def get_mod_names(mods_folder):
    """Get the names of mods in the given mods folder."""
    mod_names = []
    for mod in os.listdir(mods_folder):
        mod_path = os.path.join(mods_folder, mod)
        if os.path.isdir(mod_path):
            mod_names.append(mod)  # Use the folder name as-is, which already includes '@'
    return mod_names

def load_previous_mods_xlsx(output_folder):
    """Load the previous mod list from an Excel file if it exists."""
    previous_mods = []
    try:
        previous_file = os.path.join(output_folder, "mod_list.xlsx")
        wb = openpyxl.load_workbook(previous_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            previous_mods.append(row[0])  # Assume mod names are in the first column
    except FileNotFoundError:
        print("⚠️ Could not load previous mod list: File not found.")
    except Exception as e:
        print(f"⚠️ Could not load previous mod list: {e}")
    return previous_mods

def write_mod_list_to_excel(mod_names, previous_mods, output_path, columns, use_color):
    """Write the mod list to an Excel file."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Mods"

    # Set up headers based on columns configuration
    headers = []
    if columns.get("mod_name", False):
        headers.append("Mod Name")
    if columns.get("mod_version", False):
        headers.append("Version")
    if columns.get("status", False):
        headers.append("Status")
    sheet.append(headers)

    # Counters for Added, Removed, and Unchanged mods
    added_count = 0
    removed_count = 0
    unchanged_count = 0

    # Fill in the mod list and calculate status counts
    for mod_name in mod_names:
        row = []
        if columns.get("mod_name", False):
            row.append(mod_name)
        if columns.get("mod_version", False):
            row.append("Unknown Version")  # You can replace this with actual version extraction logic
        if columns.get("status", False):
            if mod_name in previous_mods:
                row.append("Unchanged")
                unchanged_count += 1
            else:
                row.append("New")
                added_count += 1
        sheet.append(row)

    # Auto size the columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Apply color formatting if enabled
    if use_color:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value == "New":
                    cell.fill = PatternFill(start_color="AAFFAA", end_color="AAFFAA", fill_type="solid")  # Green for new
                elif cell.value == "Unchanged":
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Light gray for unchanged

    # Save the workbook
    wb.save(output_path)
    print(f"✅ Excel file saved: {output_path}")

    # Print the counts for added, removed, and unchanged
    print(f"🟢 Added: {added_count}, 🔴 Removed: {removed_count}, ⚪ Unchanged: {unchanged_count}")

def write_mod_list_to_html(mod_names, previous_mods, output_path, columns):
    """Write the mod list to an HTML file."""
    html_content = """
    <html>
    <head>
        <title>Mod List</title>
    </head>
    <body>
        <h1>Mod List</h1>
        <table border="1">
            <tr>
                {% if 'mod_name' in columns %}<th>Mod Name</th>{% endif %}
                {% if 'mod_version' in columns %}<th>Version</th>{% endif %}
                {% if 'status' in columns %}<th>Status</th>{% endif %}
            </tr>
            {% for mod in mod_names %}
                <tr>
                    {% if 'mod_name' in columns %}<td>{{ mod }}</td>{% endif %}
                    {% if 'mod_version' in columns %}<td>Unknown Version</td>{% endif %}
                    {% if 'status' in columns %}
                        <td>{% if mod in previous_mods %}Unchanged{% else %}New{% endif %}</td>
                    {% endif %}
                </tr>
            {% endfor %}
        </table>
    </body>
    </html>
    """
    template = Template(html_content)
    rendered_html = template.render(mod_names=mod_names, previous_mods=previous_mods, columns=columns)

    with open(output_path, "w") as f:
        f.write(rendered_html)

    print(f"✅ HTML file saved: {output_path}")

def write_mod_list_to_markdown(mod_names, previous_mods, output_path, columns):
    """Write the mod list to a Markdown file."""
    markdown_content = "# Mod List\n\n"
    markdown_content += "| Mod Name | Version | Status |\n"
    markdown_content += "| --- | --- | --- |\n"

    added_count = 0
    removed_count = 0
    unchanged_count = 0

    for mod in mod_names:
        mod_version = "Unknown Version" if columns.get("mod_version", False) else ""
        status = "Unchanged" if mod in previous_mods else "New"
        if status == "Unchanged":
            unchanged_count += 1
        else:
            added_count += 1
        markdown_content += f"| {mod} | {mod_version} | {status} |\n"

    with open(output_path, "w") as f:
        f.write(markdown_content)

    print(f"✅ Markdown file saved: {output_path}")
    print(f"🟢 Added: {added_count}, 🔴 Removed: {removed_count}, ⚪ Unchanged: {unchanged_count}")
